"""
批量导入 API
支持 Excel (.xlsx) 格式导入商品、客户、供应商
"""
from flask import Blueprint, jsonify, request, send_file
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from io import BytesIO
from flask_login import login_required, current_user
from app import db
from app.models import Goods, Customer, Supplier
from app.utils.permissions import role_required
from app.api.logs import log_operation
from app.utils.scoring import score_operation
from app.utils.time_helper import beijing_now
from decimal import Decimal, InvalidOperation

bp = Blueprint('import', __name__)


def _make_template(headers, example_row, sheet_name):
    """生成 Excel 模板"""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # 表头样式
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color='E6F7FF', end_color='E6F7FF', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')

    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(1, col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # 示例数据行
    if example_row:
        ws.append(example_row)

    # 调整列宽
    for col in range(1, len(headers) + 1):
        max_len = len(str(headers[col - 1]))
        if example_row and col - 1 < len(example_row):
            max_len = max(max_len, len(str(example_row[col - 1])))
        ws.column_dimensions[ws.cell(1, col).column_letter].width = min(max_len + 4, 30)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def _parse_int(value, default=0):
    """安全解析整数"""
    if value is None:
        return default
    try:
        return int(value)
    except (ValueError, TypeError):
        return default


def _parse_decimal(value):
    """安全解析 Decimal"""
    if value is None:
        return None
    try:
        return Decimal(str(value))
    except (InvalidOperation, ValueError, TypeError):
        return None


# ==================== 模板下载 ====================

@bp.route('/import/template/goods', methods=['GET'])
@login_required
def download_goods_template():
    """下载商品导入模板"""
    headers = ['SKU', '商品名称', '规格', '单位', '最低库存', '最高库存', '采购价', '销售价']
    example = ['G001', '矿泉水', '500ml/瓶', '箱', 100, 5000, 24.00, 36.00]
    output = _make_template(headers, example, '商品导入模板')
    return send_file(
        output,
        as_attachment=True,
        download_name='商品导入模板.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@bp.route('/import/template/customers', methods=['GET'])
@login_required
def download_customers_template():
    """下载客户导入模板"""
    headers = ['客户名称', '等级', '联系人', '电话', '邮箱', '地址', '结算方式', '信用额度']
    example = ['广州贸易有限公司', 'vip', '张三', '13800138000', 'zhangsan@example.com',
               '广州市天河区', 'monthly', 50000]
    output = _make_template(headers, example, '客户导入模板')
    return send_file(
        output,
        as_attachment=True,
        download_name='客户导入模板.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@bp.route('/import/template/suppliers', methods=['GET'])
@login_required
def download_suppliers_template():
    """下载供应商导入模板"""
    headers = ['供应商名称', '联系人', '电话', '地址']
    example = ['深圳物流有限公司', '李四', '13900139000', '深圳市南山区']
    output = _make_template(headers, example, '供应商导入模板')
    return send_file(
        output,
        as_attachment=True,
        download_name='供应商导入模板.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


# ==================== 商品导入 ====================

@bp.route('/import/goods', methods=['POST'])
@role_required('admin', 'teacher', 'warehouse_keeper')
@login_required
def import_goods():
    """批量导入商品"""
    if 'file' not in request.files:
        return jsonify({'code': 400, 'message': '请上传文件', 'data': None})

    file = request.files['file']
    if not file.filename.endswith('.xlsx'):
        return jsonify({'code': 400, 'message': '仅支持 .xlsx 格式', 'data': None})

    try:
        wb = load_workbook(file, read_only=True)
        ws = wb.active
    except Exception as e:
        return jsonify({'code': 400, 'message': f'文件解析失败: {str(e)}', 'data': None})

    # 预加载已有SKU
    existing_skus = set(g.sku for g in Goods.query.filter_by(status='active').all())

    success = 0
    skipped = 0
    failed = 0
    errors = []
    rows = list(ws.iter_rows(min_row=2, values_only=True))  # 跳过表头

    for idx, row in enumerate(rows, start=2):
        # 跳过空行
        if not row or all(cell is None for cell in row):
            continue

        sku = str(row[0]).strip() if row[0] else ''
        name = str(row[1]).strip() if len(row) > 1 and row[1] else ''

        # 必填校验
        if not sku:
            errors.append({'row': idx, 'reason': 'SKU不能为空'})
            failed += 1
            continue
        if not name:
            errors.append({'row': idx, 'reason': '商品名称不能为空'})
            failed += 1
            continue

        # 唯一性校验
        if sku in existing_skus:
            errors.append({'row': idx, 'reason': f'SKU已存在: {sku}'})
            skipped += 1
            continue

        spec = str(row[2]).strip() if len(row) > 2 and row[2] else ''
        unit = str(row[3]).strip() if len(row) > 3 and row[3] else '个'
        min_stock = _parse_int(row[4], 0) if len(row) > 4 else 0
        max_stock = _parse_int(row[5], 99999) if len(row) > 5 else 99999
        purchase_price = _parse_decimal(row[6]) if len(row) > 6 else None
        selling_price = _parse_decimal(row[7]) if len(row) > 7 else None

        goods = Goods(
            sku=sku,
            name=name,
            spec=spec or None,
            unit=unit,
            min_stock=min_stock,
            max_stock=max_stock,
            purchase_price=purchase_price,
            selling_price=selling_price
        )
        db.session.add(goods)
        existing_skus.add(sku)
        success += 1

    db.session.commit()
    wb.close()

    # 操作日志 + 评分
    log_operation(
        user_id=current_user.id, group_id=current_user.group_id,
        module='goods', action='import',
        target_type='Goods', target_id=None,
        description=f'批量导入商品 成功{success}条 跳过{skipped}条 失败{failed}条'
    )
    db.session.commit()
    score_operation(user_id=current_user.id, group_id=current_user.group_id,
                    module='goods', action='import')

    return jsonify({
        'code': 200,
        'message': '导入完成',
        'data': {
            'total': success + skipped + failed,
            'success': success,
            'skipped': skipped,
            'failed': failed,
            'errors': errors
        }
    })


# ==================== 客户导入 ====================

@bp.route('/import/customers', methods=['POST'])
@role_required('admin', 'teacher')
@login_required
def import_customers():
    """批量导入客户"""
    if 'file' not in request.files:
        return jsonify({'code': 400, 'message': '请上传文件', 'data': None})

    file = request.files['file']
    if not file.filename.endswith('.xlsx'):
        return jsonify({'code': 400, 'message': '仅支持 .xlsx 格式', 'data': None})

    try:
        wb = load_workbook(file, read_only=True)
        ws = wb.active
    except Exception as e:
        return jsonify({'code': 400, 'message': f'文件解析失败: {str(e)}', 'data': None})

    valid_levels = {'vip', 'normal', 'potential'}
    valid_settlements = {'monthly', 'spot', 'credit'}

    success = 0
    skipped = 0
    failed = 0
    errors = []
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    # 生成客户编号的基础
    existing_count = Customer.query.filter_by(status='active').count()

    for idx, row in enumerate(rows, start=2):
        if not row or all(cell is None for cell in row):
            continue

        name = str(row[0]).strip() if row[0] else ''

        if not name:
            errors.append({'row': idx, 'reason': '客户名称不能为空'})
            failed += 1
            continue

        level_raw = str(row[1]).strip() if len(row) > 1 and row[1] else 'normal'
        level = level_raw if level_raw in valid_levels else 'normal'

        contact_person = str(row[2]).strip() if len(row) > 2 and row[2] else None
        phone = str(row[3]).strip() if len(row) > 3 and row[3] else None
        email = str(row[4]).strip() if len(row) > 4 and row[4] else None
        address = str(row[5]).strip() if len(row) > 5 and row[5] else None

        settlement_raw = str(row[6]).strip() if len(row) > 6 and row[6] else None
        settlement_type = settlement_raw if settlement_raw in valid_settlements else None

        credit_limit = 0.0
        if len(row) > 7 and row[7] is not None:
            try:
                credit_limit = float(row[7])
            except (ValueError, TypeError):
                credit_limit = 0.0

        existing_count += 1
        customer = Customer(
            customer_no=f'C{beijing_now().strftime("%Y%m%d")}{existing_count:04d}',
            name=name,
            level=level,
            contact_person=contact_person,
            phone=phone,
            email=email,
            address=address,
            settlement_type=settlement_type,
            credit_limit=credit_limit,
            status='active'
        )
        db.session.add(customer)
        success += 1

    db.session.commit()
    wb.close()

    log_operation(
        user_id=current_user.id, group_id=current_user.group_id,
        module='customer', action='import',
        target_type='Customer', target_id=None,
        description=f'批量导入客户 成功{success}条 跳过{skipped}条 失败{failed}条'
    )
    db.session.commit()
    score_operation(user_id=current_user.id, group_id=current_user.group_id,
                    module='customer', action='import')

    return jsonify({
        'code': 200,
        'message': '导入完成',
        'data': {
            'total': success + skipped + failed,
            'success': success,
            'skipped': skipped,
            'failed': failed,
            'errors': errors
        }
    })


# ==================== 供应商导入 ====================

@bp.route('/import/suppliers', methods=['POST'])
@role_required('admin', 'teacher')
@login_required
def import_suppliers():
    """批量导入供应商"""
    if 'file' not in request.files:
        return jsonify({'code': 400, 'message': '请上传文件', 'data': None})

    file = request.files['file']
    if not file.filename.endswith('.xlsx'):
        return jsonify({'code': 400, 'message': '仅支持 .xlsx 格式', 'data': None})

    try:
        wb = load_workbook(file, read_only=True)
        ws = wb.active
    except Exception as e:
        return jsonify({'code': 400, 'message': f'文件解析失败: {str(e)}', 'data': None})

    existing_names = set(s.name for s in Supplier.query.filter_by(status='active').all())

    success = 0
    skipped = 0
    failed = 0
    errors = []
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    for idx, row in enumerate(rows, start=2):
        if not row or all(cell is None for cell in row):
            continue

        name = str(row[0]).strip() if row[0] else ''

        if not name:
            errors.append({'row': idx, 'reason': '供应商名称不能为空'})
            failed += 1
            continue

        if name in existing_names:
            errors.append({'row': idx, 'reason': f'供应商已存在: {name}'})
            skipped += 1
            continue

        contact = str(row[1]).strip() if len(row) > 1 and row[1] else None
        phone = str(row[2]).strip() if len(row) > 2 and row[2] else None
        address = str(row[3]).strip() if len(row) > 3 and row[3] else None

        supplier = Supplier(
            name=name,
            contact=contact,
            phone=phone,
            address=address
        )
        db.session.add(supplier)
        existing_names.add(name)
        success += 1

    db.session.commit()
    wb.close()

    log_operation(
        user_id=current_user.id, group_id=current_user.group_id,
        module='supplier', action='import',
        target_type='Supplier', target_id=None,
        description=f'批量导入供应商 成功{success}条 跳过{skipped}条 失败{failed}条'
    )
    db.session.commit()
    score_operation(user_id=current_user.id, group_id=current_user.group_id,
                    module='supplier', action='import')

    return jsonify({
        'code': 200,
        'message': '导入完成',
        'data': {
            'total': success + skipped + failed,
            'success': success,
            'skipped': skipped,
            'failed': failed,
            'errors': errors
        }
    })
